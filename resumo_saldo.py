import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os


arquivos = []


def selecionar_planilhas():

    global arquivos

    arquivos = filedialog.askopenfilenames(
        title="Selecionar planilhas das obras",
        filetypes=[
            ("Excel", "*.xlsx *.xlsm")
        ]
    )

    lista.delete(0, tk.END)

    for arquivo in arquivos:
        lista.insert(
            tk.END,
            os.path.basename(arquivo)
        )



def encontrar_total(ws):

    # Procura o último TOTAL de baixo para cima
    for linha in range(ws.max_row, 0, -1):

        for coluna in range(1, ws.max_column + 1):

            valor = ws.cell(
                linha,
                coluna
            ).value


            if valor:

                texto = str(valor).strip().upper()


                if texto.startswith("TOTAL"):

                    return linha


    return None



def ler_saldos(ws):

    linha_total = encontrar_total(ws)


    if linha_total is None:

        print("TOTAL não encontrado")

        return 0, 0, 0



    # ====================================
    # COLUNAS DA PLANILHA
    #
    # N = 14
    # O = 15
    # P = 16
    #
    # ====================================


    saldo_contrato = ws.cell(
        linha_total,
        14
    ).value


    saldo_contrato_pi_r = ws.cell(
        linha_total,
        15
    ).value


    saldo_empenho = ws.cell(
        linha_total,
        16
    ).value



    print("==============================")
    print("Linha TOTAL:", linha_total)
    print("Saldo Contrato (N):", saldo_contrato)
    print("Saldo Contrato PI+R (O):", saldo_contrato_pi_r)
    print("Saldo Empenho (P):", saldo_empenho)
    print("==============================")


    return (
        saldo_contrato,
        saldo_contrato_pi_r,
        saldo_empenho
    )

from datetime import datetime

def ler_data_base(ws):

    for linha in ws.iter_rows():

        for celula in linha:

            if celula.value is None:
                continue

            texto = str(celula.value).strip().upper()

            if "DATA BASE DO ORÇAMENTO" in texto:

                # Procura o primeiro valor preenchido à direita
                for coluna in range(celula.column + 1, ws.max_column + 1):

                    valor = ws.cell(celula.row, coluna).value

                    if valor not in (None, ""):

                        if isinstance(valor, datetime):
                            return valor.strftime("%m/%Y")

                        return str(valor)

    return ""

def processar():

    if not arquivos:

        messagebox.showwarning(
            "Aviso",
            "Selecione as planilhas."
        )

        return



    resultado = []


    for arquivo in arquivos:

        try:

            wb = openpyxl.load_workbook(
                arquivo,
                data_only=True
            )


            nome_aba = "Saldo Contrato | Saldo Empenho"


            if nome_aba not in wb.sheetnames:

                print(
                    "Aba não encontrada:",
                    arquivo
                )

                continue



            ws = wb[nome_aba] 
            if "Dados da Obra" in wb.sheetnames:data_base = ler_data_base(wb["Dados da Obra"])
            
            else:data_base = ""


            contrato, contrato_pi_r, empenho = ler_saldos(ws)



            resultado.append(
                [
                    os.path.basename(arquivo),
                    contrato,
                    contrato_pi_r,
                    empenho,
                    data_base
                ]
            )


        except Exception as erro:

            print(
                "Erro:",
                arquivo,
                erro
            )



    if not resultado:

        messagebox.showwarning(
            "Aviso",
            "Nenhum resultado encontrado."
        )

        return



    salvar = filedialog.asksaveasfilename(
        title="Salvar resumo",
        defaultextension=".xlsx",
        filetypes=[
            ("Excel", "*.xlsx")
        ]
    )



    if salvar:


        wb_saida = openpyxl.Workbook()


        ws_saida = wb_saida.active

        ws_saida.title = "Resumo"



        ws_saida.append(
            [
                "Obra",
                "Saldo do Contrato",
                "Saldo do Contrato (PI + R)",
                "Saldo do Empenho",
                "Data Base do Orçamento" 
            ]
        )



        for linha in resultado:

            ws_saida.append(linha)



        # ===============================
        # FORMATAÇÃO
        # ===============================


        preenchimento = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )


        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )



        for celula in ws_saida[1]:

            celula.font = Font(
                bold=True
            )

            celula.fill = preenchimento

            celula.alignment = Alignment(
                horizontal="center"
            )



        for linha in ws_saida.iter_rows():

            for celula in linha:

                celula.border = borda



        for linha in ws_saida.iter_rows(
            min_row=2,
            min_col=2,
            max_col=5   
        ):

            for celula in linha:

                celula.number_format = '#,##0.00'



        larguras = {
            1:45,
            2:25,
            3:30,
            4:25,
            5:25
        }


        for coluna, largura in larguras.items():

            ws_saida.column_dimensions[
                get_column_letter(coluna)
            ].width = largura



        ws_saida.freeze_panes = "A2"

        ws_saida.auto_filter.ref = ws_saida.dimensions



        wb_saida.save(salvar)



        messagebox.showinfo(
            "Finalizado",
            "Resumo criado com sucesso."
        )




# ==========================
# INTERFACE
# ==========================


janela = tk.Tk()

janela.title(
    "Resumo Saldo Contrato | Saldo Empenho"
)

janela.geometry(
    "750x500"
)



tk.Button(
    janela,
    text="Selecionar Planilhas",
    command=selecionar_planilhas,
    width=30
).pack(pady=10)



lista = tk.Listbox(
    janela,
    width=100,
    height=18
)

lista.pack()



tk.Button(
    janela,
    text="Gerar Resumo",
    command=processar,
    width=30
).pack(pady=20)



janela.mainloop()